# coding:utf-8
# Author:liuyan
# time:2019/3/11 21:43
# file:do_excel.py
from openpyxl import load_workbook
from learn_api.common.config import ConfigCommon
from learn_api.common import project_path,http_request
class DoExcel:
    '''该类完成从excel中读取测试数据，写回测试结果
       file_name:测试数据的文件名
       sheet_name:表单名
    '''
    def __init__(self,file_name):
        self.file_name = file_name

    def read_data(self):
        '''读取测试数据'''
        wb = load_workbook(self.file_name)
        #从配置文件里面获取需要读取那些测试用例
        case_id = ConfigCommon().getdic('CASE','case_id')
        tel = self.get_tel()
        final_data=[]
        #{'register':'all','recharge':'all'}循环配置文件，达到获取用例数据的目的
        for sheet_name in case_id:
            sheet = wb[sheet_name]
            test_data = []
            for i in range(2,sheet.max_row +1):
                row_data={}
                row_data['CaseId'] = sheet.cell(i,1).value
                row_data['Module'] = sheet.cell(i,2).value
                row_data['url'] = sheet.cell(i, 3).value
                row_data['Title'] = sheet.cell(i, 4).value
                row_data['Method'] = sheet.cell(i, 5).value
                #实现电话号码自动加1
                if sheet.cell(i,6).value.find('tel') !=-1:
                    #替换注册用例里面的tel
                    row_data['Params'] = sheet.cell(i,6).value.replace('tel',str(tel))
                    #加1后的电话号码写入表格
                    self.write_tel(int(tel)+1)
                else:
                    row_data['Params'] = sheet.cell(i, 6).value

                row_data['ExpectedResult'] = sheet.cell(i, 7).value
                test_data.append(row_data)

            if case_id[sheet_name] == 'all':
                final_data.extend(test_data)
            else:
                for i in case_id[sheet_name]:
                    final_data.append(test_data[i-1])
            wb.close()
        return final_data
    def write_data(self,row,actualresule,testresult):
        '''将测试结果写回excel表格
            row：写回的行数
            actualresule：返回结果
            testresult：测试结论（通过还是不通过）
        '''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row,8).value = actualresule  #写入实际的返回结果
        sheet.cell(row,9).value = testresult   #写入测试结果
        wb.save(self.file_name)
        wb.close()

    def get_tel(self):
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1,2).value
    def write_tel(self,param):
        '''将新生成的电话号码写回excel
        '''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        sheet.cell(1,2).value = param
        wb.save(self.file_name)
        wb.close()
if __name__ == '__main__':
    # p = {"mobilephone":"18380461781","pwd":"123456","regname":"lyhh"}
    d = DoExcel(project_path.excel_path).read_data()

    print(d)


