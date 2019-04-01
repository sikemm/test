# coding:utf-8
# Author:liuyan
# time:2019/3/11 21:43
# file:do_excel.py
from openpyxl import load_workbook
from learn_api.common.config import ConfigCommon
from learn_api.common import project_path, http_request
from learn_api.common.do_mysql import DoMysql
from decimal import *

class DoExcel:
    '''该类完成从excel中读取测试数据，写回测试结果
       file_name:测试数据的文件名
       sheet_name:表单名
    '''

    def __init__(self, file_name,sheet_name,case_id):
        self.file_name = file_name
        self.sheet_name = sheet_name
        self.case_id = case_id
    def read_data(self):
        '''读取测试数据'''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        tel = self.get_tel()
        # case_id = ConfigCommon().getstr(section, 'case_id')

        final_data = []

        test_data = []
        for i in range(2, sheet.max_row + 1):
            row_data = {}
            row_data['CaseId'] = sheet.cell(i, 1).value
            row_data['Module'] = sheet.cell(i, 2).value
            row_data['url'] = sheet.cell(i, 3).value
            row_data['Title'] = sheet.cell(i, 4).value
            row_data['Method'] = sheet.cell(i, 5).value
                # 实现电话号码自动加1
            if sheet.cell(i, 6).value.find('tel') != -1:
                row_data['Params'] = sheet.cell(i, 6).value.replace('tel', str(tel))   # 替换注册用例里面的tel
                self.write_tel(int(tel) + 1)   # 加1后的电话号码写入表格
            else:
                row_data['Params'] = sheet.cell(i, 6).value

            if sheet.cell(i, 7).value.find('leaveAmount')!=-1:   #每次执行用例都重新查询一次客户的余额
                query = 'select LeaveAmount from member where Id = 1125548'
                leaveamount = DoMysql().do_mysql(query, 1)
                param = eval(sheet.cell(i, 6).value)
                # 查找期望值里面是否有leaveAmount，有的话就替换成数据库查找出来的值加上本次充值金额
                if sheet.cell(i, 2).value =='recharge': #如果模块名为recharge就为充值，进行加操作，否则就为取现，进行减操作
                    add_num = '%.2f'%(float(leaveamount[0]) + float(param['amount']))  #充值
                else:
                    add_num = '%.2f' % (float(leaveamount[0]) - float(param['amount']))  # 取现

                row_data['ExpectedResult'] = sheet.cell(i, 7).value.replace('leaveAmount',str(add_num))
            else:
                row_data['ExpectedResult'] = sheet.cell(i, 7).value

            row_data['sheet_name'] = self.sheet_name
            test_data.append(row_data)
        # 从配置文件里面获取需要读取那些测试用例

        if self.case_id =='all':
            final_data.extend(test_data)
                # final_data = test_data
        else:
            for i in eval(self.case_id):
                final_data.append(test_data[i - 1])
        wb.close()
        return final_data

    def write_data(self, row, actualresule, testresult):
        '''将测试结果写回excel表格
            row：写回的行数
            actualresule：返回结果
            testresult：测试结论（通过还是不通过）
        '''
        wb = load_workbook(self.file_name)
        sheet = wb[self.sheet_name]
        sheet.cell(row, 8).value = actualresule  # 写入实际的返回结果
        sheet.cell(row, 9).value = testresult  # 写入测试结果
        wb.save(self.file_name)
        wb.close()

    def get_tel(self):
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        wb.close()
        return sheet.cell(1, 2).value

    def write_tel(self, param):
        '''将新生成的电话号码写回excel
        '''
        wb = load_workbook(self.file_name)
        sheet = wb['tel']
        sheet.cell(1, 2).value = param
        wb.save(self.file_name)
        wb.close()

    def write_amount(self,row,amount):
        '''查询客户的充值余额，并写回excel
        :amount：客户余额
        '''
        wb = load_workbook(self.file_name)
        sheet = wb['recharge']
        sheet.cell(row, 7).value = amount
        wb.save(self.file_name)
        wb.close()


if __name__ == '__main__':
    # p = {"mobilephone":"18380461781","pwd":"123456","regname":"lyhh"}
    d = DoExcel(project_path.excel_path,'recharge','all').read_data()
    print(d)


